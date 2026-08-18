#!/usr/bin/env python3
"""
Gera a página estática consulta_estoque_mhc.html com os dados das duas planilhas
JÁ EMBUTIDOS no arquivo — sem tela de upload, pronta pra publicar no GitHub Pages.

Uso:
    python3 gerar_estoque_mhc.py <pasta_com_as_2_planilhas> <template.html> <saida.html>

Ele procura, dentro da pasta informada, os dois arquivos .xlsx exportados do ERP e
identifica qual é qual pelas colunas (igual a lógica já usada no próprio HTML):
  - "Saldo por endereço": tem as colunas "Endereco" e "Saldo fisico"
  - "Estoque detalhado":  tem a coluna "Obs qualidade" ou "Lote mae"

Se a pasta tiver mais de um arquivo .xlsx que bate com o mesmo papel, ou nenhum,
o script para com um erro explicando o problema (evita gerar página com dado errado).
"""
import sys
import json
import glob
import os
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERRO: falta a biblioteca 'openpyxl'. Instale com: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PLACEHOLDER = "/*MHC_EMBEDDED_DATA*/ null"


def sheet_to_rows(path):
    """Replica o comportamento de XLSX.utils.sheet_to_json(sheet, {defval: ''}) do SheetJS:
    lê a primeira aba, usa a primeira linha como cabeçalho, cada linha seguinte vira um dict
    {cabecalho: valor}, células vazias viram string vazia."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]
    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = raw_row[i] if i < len(raw_row) else None
            if v is None:
                v = ""
            elif isinstance(v, datetime):
                v = v.isoformat()
            row[h] = v
        rows.append(row)
    return rows


def classify(path):
    """Retorna 'endereco', 'detalhe' ou None, olhando os cabeçalhos da primeira linha."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    first_row = next(sheet.iter_rows(values_only=True), None)
    if not first_row:
        return None
    headers = {str(h).strip() for h in first_row if h is not None}
    if "Endereco" in headers and "Saldo fisico" in headers:
        return "endereco"
    if "Obs qualidade" in headers or "Lote mae" in headers:
        return "detalhe"
    return None


def find_sources(folder):
    candidates = sorted(glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls")))
    # ignora arquivos temporários do Excel (~$arquivo.xlsx)
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    found = {"endereco": [], "detalhe": []}
    for path in candidates:
        role = classify(path)
        if role:
            found[role].append(path)

    errors = []
    for role, label in (("endereco", "Saldo por endereço"), ("detalhe", "Estoque detalhado")):
        if len(found[role]) == 0:
            errors.append(f'Nenhum arquivo .xlsx com as colunas de "{label}" encontrado em {folder}.')
        elif len(found[role]) > 1:
            names = ", ".join(os.path.basename(p) for p in found[role])
            errors.append(
                f'Mais de um arquivo .xlsx parece ser "{label}" ({names}). '
                f'Apague o(s) arquivo(s) antigo(s) da pasta antes de gerar de novo.'
            )
    if errors:
        raise SystemExit("ERRO ao identificar as planilhas:\n- " + "\n- ".join(errors))

    return found["endereco"][0], found["detalhe"][0]


def build_embedded_json(endereco_path, detalhe_path):
    data = {
        "endereco": {
            "name": os.path.basename(endereco_path),
            "rows": sheet_to_rows(endereco_path),
        },
        "detalhe": {
            "name": os.path.basename(detalhe_path),
            "rows": sheet_to_rows(detalhe_path),
        },
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }
    return json.dumps(data, ensure_ascii=False)


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    folder, template_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    endereco_path, detalhe_path = find_sources(folder)
    print(f"Saldo por endereço: {endereco_path}")
    print(f"Estoque detalhado:  {detalhe_path}")

    embedded_json = build_embedded_json(endereco_path, detalhe_path)

    with open(template_path, "r", encoding="utf-8") as f:
        template = f.read()

    if PLACEHOLDER not in template:
        raise SystemExit(
            f'ERRO: não encontrei o marcador "{PLACEHOLDER}" no template {template_path}. '
            f'O template pode ter sido editado — confira a linha "const EMBEDDED_DATA = ...".'
        )

    output = template.replace(PLACEHOLDER, "/*MHC_EMBEDDED_DATA*/ " + embedded_json, 1)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(output)

    n_endereco = len(json.loads(embedded_json)["endereco"]["rows"])
    n_detalhe = len(json.loads(embedded_json)["detalhe"]["rows"])
    print(f"OK: {out_path} gerado com {n_endereco} linhas de endereço e {n_detalhe} linhas de detalhe.")


if __name__ == "__main__":
    main()
